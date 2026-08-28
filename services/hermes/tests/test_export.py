"""
Export helper tests.

These are the last mile: the point where a dataset stops being Parquet and
becomes a file someone opens. The failures worth catching here are all
silent ones -- a sort code that loses its leading zero, a sheet name Excel
refuses, a nested value that raises on write and takes the whole export with
it. None of them raise in the code under test unless asserted on.

Like the other tool tests these need no database and no network.
"""

from __future__ import annotations

import datetime as dt
import io
from decimal import Decimal

from openpyxl import load_workbook

from hermes.tools.report import rows_to_csv, rows_to_xlsx


def _load(payload: bytes):
    return load_workbook(io.BytesIO(payload))


def test_xlsx_round_trips_values_with_their_types():
    rows = [
        {"vendor": "ACME Ltd", "amount": 1250.5, "count": 3, "when": dt.date(2026, 8, 1)},
        {"vendor": "Globex", "amount": -150.0, "count": 0, "when": dt.date(2026, 8, 2)},
    ]

    sheet = _load(rows_to_xlsx(rows)).active

    assert [cell.value for cell in sheet[1]] == ["vendor", "amount", "count", "when"]
    # Numbers arrive as numbers, not as text Excel has to guess at.
    assert sheet["B2"].value == 1250.5
    assert sheet["C2"].value == 3
    assert sheet["B3"].value == -150.0
    assert sheet["D2"].value == dt.datetime(2026, 8, 1)


def test_xlsx_preserves_a_leading_zero_account_code():
    """
    The reason xlsx exists alongside csv.

    An account code is text that looks like a number. Handed to Excel as csv it
    becomes 41; written as a string cell it stays 0041.
    """
    sheet = _load(rows_to_xlsx([{"account": "0041"}])).active

    assert sheet["A2"].value == "0041"
    assert isinstance(sheet["A2"].value, str)


def test_xlsx_stringifies_values_openpyxl_cannot_write():
    rows = [{"meta": {"nested": True}, "tags": ["a", "b"], "total": Decimal("12.34")}]

    sheet = _load(rows_to_xlsx(rows)).active

    assert sheet["A2"].value == "{'nested': True}"
    assert sheet["B2"].value == "['a', 'b']"
    assert sheet["C2"].value == "12.34"


def test_xlsx_header_is_frozen_and_bold():
    sheet = _load(rows_to_xlsx([{"a": 1}])).active

    assert sheet.freeze_panes == "A2"
    assert sheet["A1"].font.bold is True


def test_xlsx_sheet_name_is_made_legal():
    """Excel refuses >31 characters and []:*?/\\ outright."""
    sheet = _load(rows_to_xlsx([{"a": 1}], sheet_name="Q3 sales / ACME [final]")).active
    assert sheet.title == "Q3 sales - ACME -final-"

    long_name = "a" * 60
    assert len(_load(rows_to_xlsx([{"a": 1}], sheet_name=long_name)).active.title) == 31


def test_xlsx_handles_no_rows():
    """An empty version is still a valid workbook, not a traceback."""
    sheet = _load(rows_to_xlsx([])).active

    assert sheet.max_row == 1
    assert sheet["A1"].value is None


def test_xlsx_keeps_none_empty_rather_than_writing_the_word_none():
    sheet = _load(rows_to_xlsx([{"vendor": "ACME", "note": None}])).active

    assert sheet["B2"].value is None


def test_csv_leads_with_a_utf8_bom():
    """Without it Excel mangles the pound sign, which is the first thing anyone sees."""
    payload = rows_to_csv([{"total": "£1,250.50"}])

    assert payload.startswith(b"\xef\xbb\xbf")
    assert "£1,250.50" in payload.decode("utf-8-sig")


def test_csv_of_no_rows_is_empty():
    assert rows_to_csv([]) == b""

"""
AnalyzeIt workbook parser service (Week 2 compute layer).

Implements the tool contract the dashboard's controlled tool layer expects:
parse_workbook, profile_dataset, query_dataset, apply_recipe -- plus the
webhook receiver the upload path dispatches to.

Run locally:   venv/bin/uvicorn app.main:app --host 0.0.0.0 --port 8644
"""

from __future__ import annotations

import hashlib
import io
import os
import re
from datetime import datetime
from typing import Any

import duckdb
import openpyxl
import polars as pl
from fastapi import FastAPI, Header, HTTPException, Request
from pydantic import BaseModel

from dotenv import load_dotenv
from pathlib import Path

# Load apps/web/.env.local if available (dev convenience). In a container the
# repo layout above the service dir is absent, so guard the parent lookup —
# parents[3] raises IndexError when main.py sits shallow (e.g. /app/app/main.py).
_parents = Path(__file__).resolve().parents
web_env = _parents[3] / "apps" / "web" / ".env.local" if len(_parents) > 3 else None
if web_env is not None and web_env.exists():
    load_dotenv(web_env)
else:
    load_dotenv()

APP_SECRET = os.environ.get("HERMES_WEBHOOK_SECRET") or os.environ.get("HERMES_API_SECRET", "")
TOOL_SECRET = os.environ.get("TOOL_LAYER_SECRET", "")
HERMES_API_SECRET = os.environ.get("HERMES_API_SECRET", "")


# Importing chat registers the /api/v1/chat route on this app.
# (Placed after app creation below; see bottom of module.)

# Supabase Storage (production): the webhook carries storage_path and the
# service downloads the raw workbook itself.
SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
# The deploy env (Render/render.yaml/.env) provides SUPABASE_SECRET_KEY; older
# code read SUPABASE_SERVICE_ROLE_KEY. Accept either so Storage actually works
# in production instead of silently no-opping.
SUPABASE_SERVICE_ROLE_KEY = (
    os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    or os.environ.get("SUPABASE_SECRET_KEY", "")
)
RAW_BUCKET = os.environ.get("RAW_BUCKET", "raw")
CLEANED_BUCKET = os.environ.get("CLEANED_BUCKET", "cleaned")

app = FastAPI(title="AnalyzeIt Parser Agent", version="0.1.0")

# In-memory dataset store (dataset_id -> parquet bytes + metadata).
# Swap for Supabase Storage / object store in production.
DATASETS: dict[str, dict[str, Any]] = {}


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------

def check_secret(header: str | None, expected: str, kind: str) -> None:
    valid_secrets = {
        s for s in (
            APP_SECRET,
            HERMES_API_SECRET,
            TOOL_SECRET,
            os.environ.get("HERMES_WEBHOOK_SECRET", ""),
            os.environ.get("HERMES_API_SECRET", ""),
        ) if s
    }
    if not valid_secrets:
        raise HTTPException(503, f"{kind} is not configured")
    clean = (header or "").removeprefix("Bearer ").strip()
    if clean not in valid_secrets and header not in valid_secrets:
        raise HTTPException(401, "Unauthorized")


def money(v: Any) -> float | None:
    """Parse '£1,240.00', '(150.00)', '-410.25', '1,200' -> float."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    neg = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[£$,\s]", "", s.strip("()"))
    if not s or s in {"-", "."}:
        return None
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n


def uk_date(v: Any) -> str | None:
    """Normalize dd/mm/yyyy (or datetime cells) to ISO yyyy-mm-dd."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def find_header(rows: list[list[Any]]) -> int:
    """Locate the real header row: a row where >=2 cells are non-null strings
    and at least one looks like 'Date'/'Invoice'/'Name' etc., below any title block."""
    best, best_score = 0, 0
    for i, row in enumerate(rows[:30]):
        cells = [c for c in row if c is not None]
        strings = [c for c in cells if isinstance(c, str)]
        score = len(strings)
        if score >= 2 and i > best_score * 0:  # keep simple scoring
            keywords = sum(
                1 for c in strings
                if str(c).strip().lower() in {
                    "date", "invoice", "supplier", "vendor", "net sales",
                    "vat", "amount", "description", "reference", "total",
                }
            )
            if keywords >= 2 and score > best_score:
                best, best_score = i, score
    return best


def is_junk_row(row: list[Any]) -> bool:
    first = next((c for c in row if c is not None), None)
    if isinstance(first, str):
        f = first.strip().lower()
        if f in {"subtotal", "total", "grand total"} or f.startswith(("*", "this report")):
            return True
    return all(c is None for c in row)


def normalize_vendor(name: Any) -> str:
    s = re.sub(r"\s+", " ", str(name)).strip().lower()
    s = re.sub(r"[.,]$", "", s)
    s = re.sub(r"\b(ltd|limited|ltd\.|co|company)\b", "", s).strip()
    return s


# --------------------------------------------------------------------------
# core parsing
# --------------------------------------------------------------------------

def parse_sheet(xlsx_bytes: bytes) -> dict[str, Any]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.worksheets[0]
    raw_rows = [list(r) for r in ws.iter_rows(values_only=True)]

    header_idx = find_header(raw_rows)
    columns = [str(c).strip() if c is not None else f"col_{j}"
               for j, c in enumerate(raw_rows[header_idx])]
    data_rows = raw_rows[header_idx + 1:]

    records, dropped = [], []
    for r in data_rows:
        if is_junk_row(r):
            dropped.append(r)
            continue
        rec = {}
        for j, col in enumerate(columns):
            v = r[j] if j < len(r) else None
            lc = col.lower()
            if "date" in lc:
                rec[col] = uk_date(v)
            elif any(k in lc for k in ("net", "vat", "amount", "gross", "total", "price", "value")):
                rec[col] = money(v)
            elif any(k in lc for k in ("supplier", "vendor", "name", "customer")):
                rec[col] = normalize_vendor(v) if v else None
            else:
                rec[col] = str(v).strip() if v is not None else None
        # keep only rows with at least one meaningful value beyond an invoice ref
        vals = [x for x in rec.values() if x is not None]
        if len(vals) >= 2:
            records.append(rec)

    if not records:
        df = pl.DataFrame({c: [] for c in columns})
    else:
        try:
            df = pl.DataFrame(records)
        except Exception:
            # Fallback: construct series per column to prevent schema inference crashes
            col_data = {}
            for col in columns:
                raw_vals = [rec.get(col) for rec in records]
                non_nulls = [x for x in raw_vals if x is not None]
                if non_nulls and all(isinstance(x, (int, float)) for x in non_nulls):
                    col_data[col] = pl.Series(col, [float(x) if x is not None else None for x in raw_vals], dtype=pl.Float64)
                else:
                    col_data[col] = pl.Series(col, [str(x) if x is not None else None for x in raw_vals], dtype=pl.Utf8)
            df = pl.DataFrame(col_data)
    return {
        "dataframe": df,
        "header_row": header_idx,
        "columns": columns,
        "dropped_rows": len(dropped),
        "notes": _extract_notes(wb),
    }


def _extract_notes(wb: openpyxl.Workbook) -> list[dict[str, str]]:
    """Pull Old code / New code mapping pairs from trailing sheets."""
    mappings: list[dict[str, str]] = []
    for ws in wb.worksheets[1:]:
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [c for c in row if c is not None]
            if len(cells) == 2 and all(isinstance(c, str) for c in cells):
                a, b = (str(c).strip() for c in cells)
                if i >= 2 and re.match(r"^[A-Z]{2,}[-_]?\w*$", a) and b:
                    mappings.append({"old_code": a, "new_code": b})
    return mappings


def fingerprint(df: pl.DataFrame) -> str:
    h = hashlib.sha256()
    h.update(df.columns.__str__().encode())
    h.update(df.write_csv().encode())
    return h.hexdigest()


def fetch_from_supabase(storage_path: str) -> bytes | None:
    """Download a raw upload from Supabase Storage. Returns None when the
    service is not configured for Supabase (local/dev mode)."""
    if not (SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY):
        return None
    import httpx
    url = f"{SUPABASE_URL}/storage/v1/object/{RAW_BUCKET}/{storage_path.lstrip('/')}"
    resp = httpx.get(
        url,
        headers=_supabase_headers(),
        timeout=30,
    )
    if resp.status_code != 200:
        raise HTTPException(502, f"Supabase Storage fetch failed [{resp.status_code}] for {storage_path}")
    return resp.content


def supabase_configured() -> bool:
    return bool(SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY)


def _supabase_headers(extra: dict[str, str] | None = None) -> dict[str, str]:
    # New-format keys (sb_secret_*) are not JWTs; the Storage/REST APIs require
    # them in BOTH the apikey header and Authorization: Bearer. Sending only
    # Authorization makes Storage try to parse it as a JWT ("Invalid Compact JWS").
    h = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
    }
    if extra:
        h.update(extra)
    return h


def ensure_bucket(bucket: str) -> None:
    """Create a private Storage bucket if it doesn't exist. Idempotent."""
    if not supabase_configured():
        return
    import httpx
    httpx.post(
        f"{SUPABASE_URL}/storage/v1/bucket",
        headers=_supabase_headers({"Content-Type": "application/json"}),
        json={"id": bucket, "name": bucket, "public": False},
        timeout=30,
    )  # 200 = created, 400 "already exists" = fine; we don't gate on it.


def upload_to_supabase(bucket: str, path: str, data: bytes, content_type: str) -> None:
    """Upsert bytes to Storage at bucket/path."""
    if not supabase_configured():
        raise HTTPException(409, "Supabase Storage not configured (SUPABASE_URL + service key)")
    import httpx
    url = f"{SUPABASE_URL}/storage/v1/object/{bucket}/{path.lstrip('/')}"
    resp = httpx.post(
        url,
        headers=_supabase_headers({"Content-Type": content_type, "x-upsert": "true"}),
        content=data,
        timeout=60,
    )
    if resp.status_code not in (200, 201):
        raise HTTPException(502, f"Supabase upload failed [{resp.status_code}]: {resp.text[:200]}")


def sign_supabase_url(bucket: str, path: str, expires_in: int = 3600) -> str:
    """Return a time-limited public download URL for a private object."""
    if not supabase_configured():
        raise HTTPException(409, "Supabase Storage not configured")
    import httpx
    url = f"{SUPABASE_URL}/storage/v1/object/sign/{bucket}/{path.lstrip('/')}"
    resp = httpx.post(
        url,
        headers=_supabase_headers({"Content-Type": "application/json"}),
        json={"expiresIn": expires_in},
        timeout=30,
    )
    if resp.status_code != 200:
        raise HTTPException(502, f"Supabase sign failed [{resp.status_code}]: {resp.text[:200]}")
    body = resp.json()
    signed = body.get("signedURL") or body.get("signedUrl") or ""
    return f"{SUPABASE_URL}/storage/v1{signed}"


def df_to_xlsx_bytes(df) -> bytes:
    """Serialize a Polars DataFrame to .xlsx bytes via openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(list(df.columns))
    for row in df.iter_rows():
        ws.append([v for v in row])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def ensure_parsed(dataset_id: str) -> dict[str, Any]:
    """Parse + cache a dataset on first touch (webhook or tool call)."""
    ds = DATASETS.get(dataset_id)
    if not ds and supabase_configured() and dataset_id:
        import httpx
        url = f"{SUPABASE_URL}/rest/v1/raw_uploads?or=(id.eq.{dataset_id},dataset_id.eq.{dataset_id})&status=eq.stored&limit=1"
        try:
            resp = httpx.get(url, headers=_supabase_headers(), timeout=15)
            if resp.status_code == 200 and resp.json():
                row = resp.json()[0]
                ds = {
                    "bytes": None,
                    "filename": row.get("original_filename"),
                    "storage_path": row.get("storage_path"),
                    "workspace_id": row.get("workspace_id"),
                    "dataset_id": row.get("dataset_id"),
                }
                DATASETS[dataset_id] = ds
                if row.get("id"):
                    DATASETS[row["id"]] = ds
                if row.get("dataset_id"):
                    DATASETS[row["dataset_id"]] = ds
        except Exception:
            pass

    if not ds:
        raise HTTPException(404, f"unknown dataset_id {dataset_id}")
    if "df" in ds:
        return ds
    raw = fetch_from_supabase(ds["storage_path"]) if ds.get("storage_path") else None
    if raw is None:
        raw = ds.get("bytes")
    if not raw:
        raise HTTPException(
            409,
            f"no bytes available for {dataset_id}; Supabase Storage is not configured "
            "(set SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY) or push via POST /datasets",
        )
    parsed = parse_sheet(raw)
    df = parsed["dataframe"]
    ds["df"] = df
    ds["source_signature"] = fingerprint(df)
    ds["parsed"] = {k: v for k, v in parsed.items() if k != "dataframe"}
    return ds


# --------------------------------------------------------------------------
# endpoints
# --------------------------------------------------------------------------

@app.get("/health")
def health():
    return {
        "status": "healthy",
        "agent": "AnalyzeIt Parser Agent",
        "datasets": len(DATASETS),
        "time": datetime.utcnow().isoformat(),
    }


class WebhookPayload(BaseModel):
    event: str
    dataset_id: str | None = None
    filename: str | None = None
    tenant_id: str | None = None
    workspace_id: str | None = None
    storage_path: str | None = None
    sha256: str | None = None
    instructions: str | None = None


@app.post("/webhooks/{name}")
async def webhook(name: str, payload: WebhookPayload,
                  x_hermes_secret: str | None = Header(default=None)):
    check_secret(x_hermes_secret, APP_SECRET, "HERMES_WEBHOOK_SECRET")

    result: dict[str, Any] = {"received": payload.event}
    if payload.event == "workbook.uploaded":
        ds_id = payload.dataset_id or payload.storage_path or ""
        # Register the dataset from the webhook metadata; bytes come either
        # from Supabase Storage (production) or a prior POST /datasets push.
        if ds_id and ds_id not in DATASETS:
            DATASETS[ds_id] = {
                "bytes": None, "filename": payload.filename,
                "storage_path": payload.storage_path,
            }
        try:
            ds = ensure_parsed(ds_id) if ds_id else None
        except HTTPException as exc:
            if "unknown" in str(exc.detail):
                result["warning"] = f"Bytes unavailable yet for {ds_id}; parser will fetch on first tool call"
                return result
            raise
        p = ds["parsed"]
        df = ds["df"]
        result["parse"] = {
            "rows": df.height, "columns": df.columns,
            "header_row": p["header_row"],
            "dropped_junk_rows": p["dropped_rows"],
            "mappings_found": p["notes"],
            "source_signature": ds["source_signature"],
        }
    return result


@app.post("/datasets/{dataset_id}")
async def push_dataset(dataset_id: str, request: Request,
                       x_hermes_secret: str | None = Header(default=None)):
    """Upload raw workbook bytes so webhook/tool calls can parse them."""
    check_secret(x_hermes_secret, APP_SECRET, "HERMES_WEBHOOK_SECRET")
    body = await request.body()
    if not body:
        raise HTTPException(400, "empty body")
    try:
        parse_probe = openpyxl.load_workbook(io.BytesIO(body), read_only=True)
        sheets = parse_probe.sheetnames
    except Exception:
        raise HTTPException(400, "not a valid xlsx workbook")
    raw_fn = request.headers.get("x-filename")
    import urllib.parse
    filename = urllib.parse.unquote(raw_fn) if raw_fn else dataset_id
    DATASETS[dataset_id] = {"bytes": body, "filename": filename, "sheets": sheets}
    try:
        ensure_parsed(dataset_id)
    except Exception:
        pass
    return {"stored": True, "dataset_id": dataset_id, "sheets": sheets}


@app.post("/api/v1/tools/{tool}")
async def run_tool(tool: str, request: Request,
                   authorization: str | None = Header(default=None)):
    check_secret(authorization, f"Bearer {TOOL_SECRET}" if TOOL_SECRET else "",
                 "TOOL_LAYER_SECRET")
    body = await request.json()
    params = body.get("params", {})
    dry_run = body.get("dry_run", True)

    if tool == "parse_workbook":
        ds_id = params.get("dataset_id")
        ds = ensure_parsed(ds_id)
        p = ds["parsed"]
        return {"status": "result", "result": {
            "rows": ds["df"].height, "columns": p["columns"],
            "header_row": p["header_row"],
            "dropped_junk_rows": p["dropped_rows"],
            "mappings": p["notes"],
            "source_signature": ds["source_signature"],
        }, "evidence": {"tool": tool, "dataset_id": ds_id}}

    if tool == "profile_dataset":
        ds = ensure_parsed(params.get("dataset_id"))
        df = ds["df"]
        profile = {}
        for col in df.columns:
            s = df[col]
            entry: dict[str, Any] = {
                "dtype": str(s.dtype),
                "nulls": int(s.null_count()),
                "n_unique": int(s.n_unique()),
            }
            if s.dtype.is_numeric():
                entry |= {"sum": float(s.sum()), "mean": float(s.mean()),
                          "min": float(s.min()), "max": float(s.max())}
            profile[col] = entry
        dupes = df.filter(df.is_duplicated())
        return {"status": "result", "result": {
            "rows": df.height, "duplicate_rows": dupes.height, "columns": profile},
            "evidence": {"tool": tool}}

    if tool == "query_dataset":
        ds = ensure_parsed(params.get("dataset_id"))
        sql = params.get("sql") or ""
        con = duckdb.connect()
        con.register("ds", ds["df"].to_pandas())
        rows = con.execute(sql).fetchall()
        cols = [d[0] for d in con.execute(sql).description]
        con.close()
        return {"status": "result", "result": {"columns": cols, "rows": [list(r) for r in rows]},
                "evidence": {"tool": tool, "sql": sql}}

    if tool == "apply_recipe":
        ds = ensure_parsed(params.get("dataset_id"))
        steps = params.get("steps") or []
        preview_steps = [{"step": s.get("type"), "dry_run": dry_run} for s in steps]
        if not dry_run:
            df = ds["df"]
            for step in steps:
                st = step.get("type")
                if st == "dedupe":
                    df = df.unique(keep="first")
                elif st == "drop_nulls" and step.get("column"):
                    df = df.drop_nulls(subset=[step["column"]])
            ds["cleaned"] = df
        return {"status": "result", "result": {
            "applied_steps": preview_steps if dry_run else [
                {"type": s.get("type"), "done": True} for s in steps],
            "rows_before": ds["df"].height,
            "rows_after": ds["cleaned"].height if "cleaned" in ds else None},
            "execution_metadata": {"dry_run": dry_run}}

    raise HTTPException(404, f"unknown tool '{tool}'")


# ---------------------------------------------------------------------------
# Deterministic categorize + export (no LLM, no tokens).
#
# The chat model can categorise in prose but is unreliable at driving the
# multi-step clean->export tool chain to completion. This endpoint does the
# whole job in code: add a derived label column from keyword rules, persist it
# to the cleaned copy (original never mutated), export to Supabase Storage, and
# return a signed download URL. Same building blocks the chat tools use, so the
# result is identical -- just guaranteed to finish. Used by the Hermes backend
# loop and callable directly by the dashboard.
# ---------------------------------------------------------------------------

@app.post("/api/v1/categorize")
async def categorize_and_export(request: Request,
                                authorization: str | None = Header(default=None)):
    check_secret(authorization, f"Bearer {HERMES_API_SECRET}" if HERMES_API_SECRET else "",
                 "HERMES_API_SECRET")
    body = await request.json()
    dataset_id = body.get("dataset_id")
    if not dataset_id:
        raise HTTPException(400, "dataset_id is required")

    source_column = body.get("source_column")
    target_column = body.get("target_column") or "Category"
    rules = body.get("rules") or []
    default = body.get("default", "Uncategorised")
    fmt = (body.get("format") or "xlsx").lower()

    # Reuse the verified chat-layer implementations. Imported here (not at module
    # top) because chat imports from main -- a top-level import would be circular.
    try:
        from .chat import _apply_categorize, _tool_export
    except ImportError:
        from chat import _apply_categorize, _tool_export

    ds = ensure_parsed(dataset_id)

    # If the caller did not name a source column, pick the most text-like one
    # (the description/narrative column on a bank statement). Deterministic:
    # the widest-average-length Utf8 column.
    if not source_column:
        candidates = [c for c in ds["df"].columns if str(ds["df"][c].dtype) == "Utf8"]
        if not candidates:
            candidates = list(ds["df"].columns)
        best, best_len = None, -1.0
        for c in candidates:
            try:
                avg = ds["df"][c].cast(pl_str()).str.len_chars().mean() or 0.0
            except Exception:
                avg = 0.0
            if avg > best_len:
                best, best_len = c, avg
        source_column = best

    if not rules:
        raise HTTPException(400, "rules are required (list of {category, keywords})")

    step = {
        "type": "categorize",
        "source_column": source_column,
        "target_column": target_column,
        "rules": rules,
        "default": default,
    }
    working, note = _apply_categorize(ds["df"], step)
    if "skipped" in note:
        raise HTTPException(400, f"categorize skipped: {note['skipped']}")

    ds["cleaned"] = working
    export = _tool_export(dataset_id, fmt, "cleaned")
    if isinstance(export, dict) and export.get("error"):
        raise HTTPException(503, export["error"])

    return {
        "status": "result",
        "result": {
            "categorized": note,
            "download_url": export.get("download_url"),
            "filename": f"{export.get('exported', 'cleaned')}.{export.get('format', fmt)}",
            "format": export.get("format", fmt),
            "rows": export.get("rows"),
            "bucket": export.get("bucket"),
            "path": export.get("path"),
            "expires_in_seconds": export.get("expires_in_seconds"),
        },
        "evidence": {"tool": "categorize_and_export", "dataset_id": dataset_id,
                     "source_column": source_column},
    }


def pl_str():
    import polars as pl
    return pl.Utf8


# Import chat so its decorators register /api/v1/chat and extended /health on this app instance
try:
    from . import chat  # noqa: E402,F401
except ImportError:
    import chat  # noqa: E402,F401



if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8100)

